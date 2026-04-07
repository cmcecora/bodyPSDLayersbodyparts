import openpyxl
import json
import glob
import os

# Load the workbook
wb = openpyxl.load_workbook('docs/icd10cm_codes_2026.xlsx')
ws = wb['icd10cm_codes_2026']

# Add header for column C if not present
if ws.cell(row=1, column=3).value is None:
    ws.cell(row=1, column=3, value='Body Parts')

# Find all result files
result_files = sorted(glob.glob('docs/body_parts_results/results_*.json'))
print(f"Found {len(result_files)} result files")

total_filled = 0
for rf in result_files:
    with open(rf, 'r') as f:
        results = json.load(f)
    
    for entry in results:
        row_num = entry['row_num']
        body_parts = entry.get('body_parts', '')
        if body_parts:
            ws.cell(row=row_num, column=3, value=body_parts)
            total_filled += 1

print(f"Filled {total_filled} rows with body parts")

# Save
wb.save('docs/icd10cm_codes_2026.xlsx')
print("Saved updated Excel file")
wb.close()
